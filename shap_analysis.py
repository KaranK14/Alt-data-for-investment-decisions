"""
SHAP Analysis Script for Clinical Trial Feasibility Prediction Models

This script loads a saved model and performs SHAP analysis to understand
feature importance and model interpretability.

Usage:
    python scripts/shap_analysis.py
"""

import pandas as pd
import numpy as np
import joblib
import json
import shap
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import sys
import warnings
from typing import Tuple, Dict, List, Optional
from datetime import datetime

# Try to import openpyxl for Excel export, fallback to xlsxwriter
try:
    import openpyxl
    EXCEL_ENGINE = 'openpyxl'
except ImportError:
    try:
        import xlsxwriter
        EXCEL_ENGINE = 'xlsxwriter'
    except ImportError:
        EXCEL_ENGINE = None
        print("WARNING: Neither openpyxl nor xlsxwriter installed. Excel export will be disabled.")

# Add parent directory to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.config import (
    DATASET_PATH,
    LABEL_COL,
    ID_COL,
    RANDOM_SEED,
    MODELS_DIR,
    RESULTS_DIR,
)
from scripts.data_loading import (
    load_dataset,
    split_features_labels,
    train_val_test_split,
    identify_tabular_columns,
    identify_pca_columns,
    remove_constant_features,
)
from scripts.feature_selection import select_features_with_xgboost, apply_feature_mask
from scripts.models.xgb_model import convert_to_numeric

# Suppress warnings
warnings.filterwarnings('ignore')

# SHAP output directory
SHAP_DIR = RESULTS_DIR / "shap"
SHAP_DIR.mkdir(parents=True, exist_ok=True)

# Known leakage columns to check
LEAKAGE_COLUMNS_CHECK = [
    "completion_date",
    "primary_completion_date",
    "trial_duration_days",
    "why_stopped",
    "enrollment_actual_num",
    "overall_status",
    "last_known_status",
    "has_results",
    "is_historical",
]


def load_model_and_metadata() -> Tuple[any, Dict, Optional[any]]:
    """Load the saved model, metadata, and scaler if available."""
    model_path = MODELS_DIR / "best_model.pkl"
    metadata_path = MODELS_DIR / "best_model_metadata.json"
    scaler_path = MODELS_DIR / "best_model_scaler.pkl"
    
    if not model_path.exists():
        raise FileNotFoundError(f"Model not found: {model_path}")
    if not metadata_path.exists():
        raise FileNotFoundError(f"Metadata not found: {metadata_path}")
    
    print(f"Loading model from: {model_path}")
    model = joblib.load(model_path)
    
    print(f"Loading metadata from: {metadata_path}")
    with open(metadata_path, 'r') as f:
        metadata = json.load(f)
    
    scaler = None
    if scaler_path.exists():
        print(f"Loading scaler from: {scaler_path}")
        scaler = joblib.load(scaler_path)
    
    return model, metadata, scaler


def reconstruct_features(df: pd.DataFrame, metadata: Dict) -> pd.DataFrame:
    """
    Reconstruct the exact feature matrix used during training.
    
    This applies the same preprocessing steps:
    1. Select appropriate columns (tabular vs tabular+embeddings)
    2. Remove constant features
    3. Apply feature selection (if enabled)
    """
    print("\n" + "=" * 80)
    print("Reconstructing Feature Matrix")
    print("=" * 80)
    
    input_mode = metadata['input_mode']
    
    # Split features based on mode
    if input_mode == 'tabular_plus_embeddings':
        X, y = split_features_labels(df, mode="tabular_plus_embeddings")
        print(f"  Mode: tabular_plus_embeddings")
    else:
        X, y = split_features_labels(df, mode="tabular")
        print(f"  Mode: tabular")
    
    print(f"  Initial features: {X.shape[1]}")
    
    # Remove constant features (same as training)
    X = remove_constant_features(X)
    print(f"  After removing constants: {X.shape[1]}")
    
    # Note: Feature selection was applied during training, but we need to
    # match the exact features from metadata
    expected_features = metadata.get('feature_names', [])
    
    if expected_features:
        # Filter to only features that exist in both
        available_features = [f for f in expected_features if f in X.columns]
        missing_features = [f for f in expected_features if f not in X.columns]
        
        if missing_features:
            print(f"  WARNING: {len(missing_features)} features missing from dataset:")
            print(f"    {missing_features[:10]}..." if len(missing_features) > 10 else f"    {missing_features}")
        
        X = X[available_features]
        print(f"  Final features: {X.shape[1]} (expected: {len(expected_features)})")
    
    return X, y


def apply_preprocessing(X: pd.DataFrame, metadata: Dict, scaler: Optional[any] = None) -> Tuple[np.ndarray, List[str]]:
    """
    Apply the same preprocessing as during training.
    
    Returns:
        Transformed feature matrix and feature names
    """
    model_name = metadata['model_name']
    
    if model_name == 'XGBoost':
        # Convert to numeric (same as training)
        X_processed = convert_to_numeric(X)
        feature_names = list(X_processed.columns)
        X_array = X_processed.values
        
    elif model_name == 'RandomForest':
        # Convert to numeric, fillna
        X_processed = X.select_dtypes(include=[np.number]).fillna(0)
        feature_names = list(X_processed.columns)
        X_array = X_processed.values
        
    elif model_name == 'SVM':
        # Convert to numeric, fillna, then scale
        X_processed = X.select_dtypes(include=[np.number]).fillna(0)
        feature_names = list(X_processed.columns)
        
        if scaler is not None:
            X_array = scaler.transform(X_processed)
        else:
            print("  WARNING: No scaler found for SVM model, using raw features")
            X_array = X_processed.values
    else:
        # Default: convert to numeric
        X_processed = X.select_dtypes(include=[np.number]).fillna(0)
        feature_names = list(X_processed.columns)
        X_array = X_processed.values
    
    return X_array, feature_names


def check_leakage(feature_names: List[str]) -> bool:
    """Check if any leakage columns are present in features."""
    found_leakage = []
    for col in LEAKAGE_COLUMNS_CHECK:
        # Check exact match and partial matches
        for feat in feature_names:
            if col.lower() in feat.lower() or feat.lower() in col.lower():
                found_leakage.append(feat)
                break
    
    if found_leakage:
        print("\n" + "!" * 80)
        print("LEAKAGE DETECTED!")
        print("!" * 80)
        print("The following leakage columns were found in features:")
        for feat in found_leakage:
            print(f"  - {feat}")
        print("\nThis is a CRITICAL ERROR. Exiting.")
        return True
    
    print("\n✓ Leakage check passed: No leakage columns found in features")
    return False


def group_onehot_features(shap_df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate one-hot encoded features back to their base feature names.
    
    Example: phase_simple_Phase2, phase_simple_Phase3 -> phase_simple
    
    Expects shap_df to have feature names as index and 'mean_abs_shap' as column.
    Excludes PCA features (pca_emb_*) as they are not interpretable.
    """
    grouped_dict = {}
    feature_names = shap_df.index.tolist()
    
    # Exclude PCA features from grouping (they're not interpretable)
    feature_names = [f for f in feature_names if not f.startswith('pca_emb')]
    
    for col in feature_names:
        # Ensure col is a string
        if not isinstance(col, str):
            col = str(col)
        
        # Check if this looks like a one-hot encoded feature
        # Pattern: base_feature_category
        parts = col.split('_')
        if len(parts) >= 2:
            # Try to find base feature (everything except last part)
            # But be careful - some features might legitimately have underscores
            # Heuristic: if we see multiple features with same prefix, group them
            base_candidates = ['_'.join(parts[:-1]), '_'.join(parts[:-2]) if len(parts) >= 3 else None]
            
            grouped = False
            for base in base_candidates:
                if base and base != col:
                    # Check if there are other features with this base
                    matching = [c for c in feature_names if isinstance(c, str) and c.startswith(base + '_')]
                    if len(matching) > 1:
                        # Group these
                        if base not in grouped_dict:
                            grouped_dict[base] = []
                        grouped_dict[base].append(col)
                        grouped = True
                        break
            
            if not grouped:
                # No grouping found, keep as is
                if col not in [item for sublist in grouped_dict.values() for item in sublist]:
                    grouped_dict[col] = [col]
        else:
            # Single part, keep as is
            if col not in [item for sublist in grouped_dict.values() for item in sublist]:
                grouped_dict[col] = [col]
    
    # Aggregate SHAP values for grouped features
    grouped_shap = {}
    for base, features in grouped_dict.items():
        if len(features) > 1:
            # Sum absolute values for grouped features
            grouped_shap[base] = shap_df.loc[features, 'mean_abs_shap'].sum()
        else:
            grouped_shap[base] = shap_df.loc[features[0], 'mean_abs_shap']
    
    grouped_df = pd.DataFrame({
        'mean_abs_shap': grouped_shap
    }).sort_values('mean_abs_shap', ascending=False)
    
    # Ensure index has a name for easier reset_index() later
    grouped_df.index.name = 'feature'
    
    return grouped_df


def compute_shap_values(
    model: any,
    X_test: np.ndarray,
    feature_names: List[str],
    metadata: Dict,
    max_samples: int = 2000
) -> Tuple[np.ndarray, shap.Explainer]:
    """Compute SHAP values for the test set."""
    print("\n" + "=" * 80)
    print("Computing SHAP Values")
    print("=" * 80)
    
    model_name = metadata['model_name']
    
    # Sample if needed
    if len(X_test) > max_samples:
        print(f"  Sampling {max_samples} rows from {len(X_test)} test samples")
        np.random.seed(42)
        sample_idx = np.random.choice(len(X_test), max_samples, replace=False)
        X_shap = X_test[sample_idx]
    else:
        X_shap = X_test
        sample_idx = np.arange(len(X_test))
    
    print(f"  Computing SHAP for {len(X_shap)} samples")
    print(f"  Model type: {model_name}")
    
    # Select appropriate explainer
    if model_name in ['XGBoost', 'RandomForest']:
        print("  Using TreeExplainer")
        explainer = shap.TreeExplainer(model)
        shap_values = explainer.shap_values(X_shap)
        
        # TreeExplainer returns list for binary classification, take class 1
        if isinstance(shap_values, list):
            shap_values = shap_values[1]
    
    elif model_name == 'SVM':
        print("  Using LinearExplainer (kernel approximation)")
        # For SVM with RBF kernel, use KernelExplainer with a sample
        # But LinearExplainer works if we approximate
        background = shap.sample(X_shap, 100)  # Background sample
        explainer = shap.KernelExplainer(model.predict_proba, background)
        shap_values = explainer.shap_values(X_shap, nsamples=100)
        
        # Take class 1 probabilities
        if isinstance(shap_values, list):
            shap_values = shap_values[1]
    
    else:
        print(f"  Using KernelExplainer (fallback for {model_name})")
        background = shap.sample(X_shap, min(100, len(X_shap)))
        explainer = shap.KernelExplainer(model.predict_proba, background)
        shap_values = explainer.shap_values(X_shap, nsamples=100)
        
        if isinstance(shap_values, list):
            shap_values = shap_values[1]
    
    print(f"  SHAP values shape: {shap_values.shape}")
    
    return shap_values, explainer, sample_idx


def format_feature_name(feature_name: str) -> str:
    """Format feature name for display: replace underscores with spaces and title case."""
    return feature_name.replace('_', ' ').title()


def generate_custom_shap_bar_plot(
    importance_df: pd.DataFrame,
    output_dir: Path,
    top_n: int = 15,
    figsize: Tuple[int, int] = (10, 8)
):
    """Generate a custom high-quality bar plot of SHAP feature importance."""
    print(f"\n  Generating custom SHAP bar plot (top {top_n} features)...")
    
    # Get top N features
    top_features = importance_df.head(top_n).copy()
    
    # Sort by value (ascending for horizontal bar plot)
    top_features = top_features.sort_values('mean_abs_shap', ascending=True)
    
    # Create figure with larger size
    fig, ax = plt.subplots(figsize=figsize)
    
    # Create horizontal bar plot
    y_pos = np.arange(len(top_features))
    bars = ax.barh(y_pos, top_features['mean_abs_shap'], 
                   color='steelblue', alpha=0.8, edgecolor='black', linewidth=0.5)
    
    # Format feature names (replace underscores, title case)
    feature_labels = [format_feature_name(feat) for feat in top_features.index]
    
    # Set y-axis labels with larger font
    ax.set_yticks(y_pos)
    ax.set_yticklabels(feature_labels, fontsize=16, fontweight='medium')
    
    # Set x-axis label with larger font
    ax.set_xlabel('Mean(|SHAP|) (average impact on model output magnitude)', 
                  fontsize=14, fontweight='bold')
    
    # Set title with larger font
    ax.set_title(f'Top {top_n} Features by SHAP Importance\nDrivers of Completion Probability', 
                 fontsize=16, fontweight='bold', pad=15)
    
    # Add grid for better readability
    ax.grid(True, alpha=0.3, axis='x', linestyle='--')
    ax.set_axisbelow(True)
    
    # Add value labels on bars (show 8 decimal places for accuracy)
    max_value = top_features['mean_abs_shap'].max()
    label_offset = max_value * 0.02  # 2% of max value as offset
    for i, (idx, row) in enumerate(top_features.iterrows()):
        value = row['mean_abs_shap']
        # Show 8 decimal places for full precision
        ax.text(value + label_offset, i, f'{value:.8f}', 
                va='center', fontsize=11, fontweight='medium')
    
    # Invert y-axis so highest value is at top
    ax.invert_yaxis()
    
    # Adjust layout
    plt.tight_layout()
    
    # Save high-resolution figure
    output_path = output_dir / f"shap_top{top_n}_bar_custom.png"
    plt.savefig(output_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f"    Saved: {output_path.name}")


def generate_shap_plots(
    shap_values: np.ndarray,
    X_shap: np.ndarray,
    feature_names: List[str],
    importance_df: pd.DataFrame,
    output_dir: Path
):
    """Generate SHAP summary plots (excluding PCA features for interpretability)."""
    print("\n" + "=" * 80)
    print("Generating SHAP Plots (Non-PCA Features Only)")
    print("=" * 80)
    
    # Filter out PCA features for interpretability
    non_pca_mask = ~np.array([f.startswith('pca_emb') for f in feature_names])
    non_pca_indices = np.where(non_pca_mask)[0]
    
    # Filter SHAP values, X_shap, and feature names to non-PCA only
    shap_values_non_pca = shap_values[:, non_pca_indices]
    X_shap_non_pca = X_shap[:, non_pca_indices]
    feature_names_non_pca = [feature_names[i] for i in non_pca_indices]
    
    print(f"  Total features: {len(feature_names)}")
    print(f"  Non-PCA features: {len(feature_names_non_pca)}")
    print(f"  PCA features excluded: {len(feature_names) - len(feature_names_non_pca)}")
    
    # Replace underscores with spaces in feature names for better readability
    feature_names_display = [format_feature_name(name) for name in feature_names_non_pca]
    
    # Set larger font sizes for SHAP plots
    plt.rcParams.update({'font.size': 14, 'axes.labelsize': 14, 'axes.titlesize': 16, 
                         'xtick.labelsize': 12, 'ytick.labelsize': 14})
    
    # Create SHAP summary object (bar plot) - non-PCA features only
    shap.summary_plot(
        shap_values_non_pca,
        X_shap_non_pca,
        feature_names=feature_names_display,
        show=False,
        plot_type="bar",
        max_display=20  # Show top 20 features
    )
    plt.tight_layout()
    plt.savefig(output_dir / "shap_summary_bar.png", dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  Saved: shap_summary_bar.png (non-PCA features only)")
    
    # Beeswarm plot - non-PCA features only
    shap.summary_plot(
        shap_values_non_pca,
        X_shap_non_pca,
        feature_names=feature_names_display,
        show=False,
        plot_type="dot",
        max_display=20  # Show top 20 features
    )
    plt.tight_layout()
    plt.savefig(output_dir / "shap_summary_beeswarm.png", dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  Saved: shap_summary_beeswarm.png (non-PCA features only)")
    
    # Filter importance_df to exclude PCA features for custom plots
    importance_df_non_pca = importance_df[~importance_df.index.str.startswith('pca_emb')]
    
    # Generate custom bar plot (already excludes PCA if importance_df is filtered)
    generate_custom_shap_bar_plot(importance_df_non_pca, output_dir, top_n=15)
    generate_custom_shap_bar_plot(importance_df_non_pca, output_dir, top_n=10)
    
    # Reset font size to default
    plt.rcParams.update(plt.rcParamsDefault)


def generate_dependence_plots(
    shap_values: np.ndarray,
    X_shap: np.ndarray,
    feature_names: List[str],
    top_features: List[str],
    output_dir: Path
):
    """Generate SHAP dependence plots for top features (non-PCA only)."""
    print("\n  Generating dependence plots for top features (non-PCA only)...")
    
    # Filter out PCA features for interpretability
    top_features = [f for f in top_features if not f.startswith('pca_emb')]
    
    if not top_features:
        print("  Skipping: all requested features are PCA (excluded for interpretability)")
        return
    
    for i, feature in enumerate(top_features[:5]):
        # Find matching features in original feature list
        # Handle grouped features (e.g., "pca_emb" should match "pca_emb_0", "pca_emb_1", etc.)
        matching_features = [f for f in feature_names if feature in f or f.startswith(feature + '_')]
        
        if not matching_features:
            print(f"    WARNING: Feature {feature} not found, skipping")
            continue
        
        # Use the first matching feature (or the one that exactly matches)
        if feature in feature_names:
            feat_name = feature
        else:
            feat_name = matching_features[0]  # Use first match
        
        feat_idx = feature_names.index(feat_name)
        
        # Create safe filename (replace special characters)
        safe_feature_name = feature.replace('/', '_').replace('\\', '_').replace(' ', '_')
        
        shap.dependence_plot(
            feat_idx,
            shap_values,
            X_shap,
            feature_names=feature_names,
            show=False
        )
        plt.tight_layout()
        plt.savefig(output_dir / f"shap_dependence_{safe_feature_name}.png", dpi=300, bbox_inches='tight')
        plt.close()
        print(f"    Saved: shap_dependence_{safe_feature_name}.png")


def compute_feature_importance(shap_values: np.ndarray, feature_names: List[str]) -> pd.DataFrame:
    """Compute mean absolute SHAP values per feature."""
    mean_abs_shap = np.abs(shap_values).mean(axis=0)
    
    importance_df = pd.DataFrame({
        'feature': feature_names,
        'mean_abs_shap': mean_abs_shap
    }).sort_values('mean_abs_shap', ascending=False)
    
    # Set feature as index for easier processing in group_onehot_features
    importance_df = importance_df.set_index('feature')
    # Ensure index has a name for consistency when reset_index() is called later
    importance_df.index.name = 'feature'
    
    return importance_df


def export_to_excel(
    importance_df: pd.DataFrame,
    grouped_df: pd.DataFrame,
    output_dir: Path
):
    """Export SHAP results to Excel with detailed feature names and values."""
    if EXCEL_ENGINE is None:
        print("  WARNING: Excel export skipped (no Excel engine available)")
        return
    
    print("\n" + "=" * 80)
    print("Exporting SHAP Results to Excel")
    print("=" * 80)
    
    excel_path = output_dir / "shap_feature_importance.xlsx"
    
    try:
        with pd.ExcelWriter(excel_path, engine=EXCEL_ENGINE) as writer:
            # Sheet 1: All features (raw)
            df_all = importance_df.reset_index()
            df_all['rank'] = range(1, len(df_all) + 1)
            df_all['feature_display'] = df_all['feature'].apply(format_feature_name)
            df_all = df_all[['rank', 'feature', 'feature_display', 'mean_abs_shap']]
            df_all.to_excel(writer, sheet_name='All Features', index=False)
            
            # Sheet 2: Top 15 features
            df_top15 = importance_df.head(15).reset_index()
            df_top15['rank'] = range(1, len(df_top15) + 1)
            df_top15['feature_display'] = df_top15['feature'].apply(format_feature_name)
            df_top15 = df_top15[['rank', 'feature', 'feature_display', 'mean_abs_shap']]
            df_top15.to_excel(writer, sheet_name='Top 15 Features', index=False)
            
            # Sheet 3: Top 10 features
            df_top10 = importance_df.head(10).reset_index()
            df_top10['rank'] = range(1, len(df_top10) + 1)
            df_top10['feature_display'] = df_top10['feature'].apply(format_feature_name)
            df_top10 = df_top10[['rank', 'feature', 'feature_display', 'mean_abs_shap']]
            df_top10.to_excel(writer, sheet_name='Top 10 Features', index=False)
            
            # Sheet 4: Grouped features (if different from raw)
            if len(grouped_df) != len(importance_df):
                df_grouped = grouped_df.reset_index()
                df_grouped['rank'] = range(1, len(df_grouped) + 1)
                df_grouped['feature_display'] = df_grouped['feature'].apply(format_feature_name)
                df_grouped = df_grouped[['rank', 'feature', 'feature_display', 'mean_abs_shap']]
                df_grouped.to_excel(writer, sheet_name='Grouped Features', index=False)
            
            # Format worksheets (if using openpyxl)
            if EXCEL_ENGINE == 'openpyxl':
                workbook = writer.book
                for sheet_name in writer.sheetnames:
                    worksheet = workbook[sheet_name]
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Format header row
                    header_fill = openpyxl.styles.PatternFill(start_color="366092", 
                                                           end_color="366092", fill_type="solid")
                    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    # Format mean_abs_shap column as number with 8 decimal places
                    if 'mean_abs_shap' in [cell.value for cell in worksheet[1]]:
                        shap_col_idx = None
                        for idx, cell in enumerate(worksheet[1], 1):
                            if cell.value == 'mean_abs_shap':
                                shap_col_idx = idx
                                break
                        
                        if shap_col_idx:
                            # Format as number with 8 decimal places for full precision
                            for row in worksheet.iter_rows(min_row=2, min_col=shap_col_idx, max_col=shap_col_idx):
                                for cell in row:
                                    if cell.value is not None:
                                        cell.number_format = '0.00000000'
        
        print(f"  Saved: {excel_path.name}")
        print(f"    Sheets: {', '.join(writer.sheetnames)}")
        
    except Exception as e:
        print(f"  ERROR: Failed to export to Excel: {str(e)}")
        print(f"  Falling back to CSV export...")
        
        # Fallback to CSV (with full precision)
        df_all = importance_df.reset_index()
        df_all['feature_display'] = df_all['feature'].apply(format_feature_name)
        df_all = df_all[['feature', 'feature_display', 'mean_abs_shap']]
        df_all.to_csv(output_dir / "shap_all_features.csv", index=False, float_format='%.8f')
        
        df_top15 = importance_df.head(15).reset_index()
        df_top15['rank'] = range(1, len(df_top15) + 1)
        df_top15['feature_display'] = df_top15['feature'].apply(format_feature_name)
        df_top15 = df_top15[['rank', 'feature', 'feature_display', 'mean_abs_shap']]
        df_top15.to_csv(output_dir / "shap_top15_features.csv", index=False, float_format='%.8f')
        
        df_top10 = importance_df.head(10).reset_index()
        df_top10['rank'] = range(1, len(df_top10) + 1)
        df_top10['feature_display'] = df_top10['feature'].apply(format_feature_name)
        df_top10 = df_top10[['rank', 'feature', 'feature_display', 'mean_abs_shap']]
        df_top10.to_csv(output_dir / "shap_top10_features.csv", index=False, float_format='%.8f')
        print(f"  Saved CSV files as fallback (with full precision)")


def generate_report(
    metadata: Dict,
    importance_df: pd.DataFrame,
    grouped_df: pd.DataFrame,
    X_shap: np.ndarray,
    shap_values: np.ndarray,
    feature_names: List[str],
    output_dir: Path
):
    """Generate SHAP interpretation report."""
    print("\n" + "=" * 80)
    print("Generating SHAP Report")
    print("=" * 80)
    
    report_path = output_dir / "SHAP_REPORT.md"
    
    # Check for leakage
    leakage_found = False
    leakage_features = []
    for col in LEAKAGE_COLUMNS_CHECK:
        matching = [f for f in feature_names if col.lower() in f.lower() or f.lower() in col.lower()]
        if matching:
            leakage_found = True
            leakage_features.extend(matching)
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("# SHAP Analysis Report\n\n")
        f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        f.write("## Model Information\n\n")
        f.write(f"- **Model Type:** {metadata['model_name']}\n")
        f.write(f"- **Input Mode:** {metadata['input_mode']}\n")
        f.write(f"- **Imbalance Strategy:** {metadata['imbalance_strategy']}\n")
        f.write(f"- **ROC-AUC:** {metadata['roc_auc']:.4f} ± {metadata['roc_auc_std']:.4f}\n")
        # Count non-PCA features for report
        non_pca_count = len([f for f in feature_names if not f.startswith('pca_emb')])
        f.write(f"- **Total Features:** {len(feature_names)}\n")
        f.write(f"- **Non-PCA Features (analyzed):** {non_pca_count}\n")
        f.write(f"- **PCA Features (excluded):** {len(feature_names) - non_pca_count}\n")
        f.write(f"- **SHAP Samples:** {len(X_shap)}\n\n")
        
        f.write("## Top 10 Features by Mean(|SHAP|) - Grouped (Non-PCA Features Only)\n\n")
        f.write("**Note**: PCA embedding features (pca_emb_*) are excluded from all plots and analysis for interpretability. PCA features are not directly interpretable as they represent compressed text embeddings.\n\n")
        f.write("| Rank | Feature | Mean(|SHAP|) |\n")
        f.write("|------|---------|---------------|\n")
        for i, (feat, row) in enumerate(grouped_df.head(10).iterrows(), 1):
            f.write(f"| {i} | {feat} | {row['mean_abs_shap']:.6f} |\n")
        f.write("\n")
        
        f.write("## Top 5 Feature Interpretations\n\n")
        top5_features = grouped_df.head(5).index.tolist()
        
        for feature in top5_features:
            # Find corresponding features in original list
            matching_features = [f for f in feature_names if feature in f or f in feature]
            if not matching_features:
                continue
            
            # Get feature indices
            feat_indices = [feature_names.index(f) for f in matching_features if f in feature_names]
            if not feat_indices:
                continue
            
            # Compute correlation between feature value and SHAP value
            # Aggregate if multiple features match
            feature_values = X_shap[:, feat_indices].mean(axis=1) if len(feat_indices) > 1 else X_shap[:, feat_indices[0]]
            shap_vals = shap_values[:, feat_indices].mean(axis=1) if len(feat_indices) > 1 else shap_values[:, feat_indices[0]]
            
            correlation = np.corrcoef(feature_values, shap_vals)[0, 1]
            
            direction = "increases" if correlation > 0 else "decreases"
            f.write(f"### {feature}\n\n")
            f.write(f"- **Mean(|SHAP|):** {grouped_df.loc[feature, 'mean_abs_shap']:.6f}\n")
            f.write(f"- **Correlation (feature value <-> SHAP):** {correlation:.4f}\n")
            f.write(f"- **Interpretation:** Higher values of {feature} {direction} the predicted probability of trial completion.\n\n")
        
        f.write("## Leakage Sanity Check\n\n")
        if leakage_found:
            f.write("**⚠️ WARNING: LEAKAGE DETECTED!**\n\n")
            f.write("The following leakage columns were found in features:\n")
            for col in LEAKAGE_COLUMNS_CHECK:
                matching = [f for f in feature_names if col.lower() in f.lower() or f.lower() in col.lower()]
                if matching:
                    f.write(f"- {col}: **FOUND** ({', '.join(matching)})\n")
                else:
                    f.write(f"- {col}: NOT FOUND\n")
        else:
            f.write("**✓ Leakage Check Passed**\n\n")
            f.write("All known leakage columns were checked and NOT FOUND in features:\n")
            for col in LEAKAGE_COLUMNS_CHECK:
                f.write(f"- {col}: NOT FOUND\n")
        f.write("\n")
        
        f.write("## Notes\n\n")
        f.write("- SHAP values are computed on the test set only.\n")
        f.write("- **PCA embedding features (pca_emb_*) are excluded from all plots and analysis for interpretability.**\n")
        f.write("- PCA features are not directly interpretable individually (they represent compressed text embeddings).\n")
        f.write("- Feature importance is measured as mean absolute SHAP value across all test samples.\n")
        f.write("- Maximum of 2000 test samples used for SHAP computation (sampled with random_state=42).\n\n")
        
        f.write("## Generated Files\n\n")
        f.write("- `shap_summary_bar.png`: Global feature importance (bar plot)\n")
        f.write("- `shap_summary_beeswarm.png`: SHAP value distribution (beeswarm plot)\n")
        f.write("- `shap_top20_raw.csv`: Top 20 features with raw feature names\n")
        f.write("- `shap_top20_grouped.csv`: Top 20 features with one-hot features grouped\n")
        f.write("- `shap_dependence_<feature>.png`: Dependence plots for top 5 features\n")
        f.write("- `transformed_feature_names.txt`: List of all feature names used\n\n")
    
    print(f"  Saved: SHAP_REPORT.md")


def main():
    """Main SHAP analysis pipeline."""
    print("=" * 80)
    print("SHAP Analysis for Clinical Trial Feasibility Prediction")
    print("=" * 80)
    print(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    try:
        # 1. Load model and metadata
        model, metadata, scaler = load_model_and_metadata()
        print(f"\nModel loaded: {metadata['model_name']}")
        print(f"Input mode: {metadata['input_mode']}")
        print(f"Imbalance strategy: {metadata['imbalance_strategy']}")
        
        # 2. Load dataset
        print("\n" + "=" * 80)
        print("Loading Dataset")
        print("=" * 80)
        df = load_dataset()
        print(f"Dataset shape: {df.shape[0]:,} rows × {df.shape[1]} columns")
        
        # 3. Reconstruct features
        X, y = reconstruct_features(df, metadata)
        
        # 4. Apply same train/test split
        print("\n" + "=" * 80)
        print("Applying Train/Test Split")
        print("=" * 80)
        X_train, X_val, X_test, y_train, y_val, y_test = train_val_test_split(
            X, y, random_state=RANDOM_SEED
        )
        print(f"  Train: {len(X_train):,} samples")
        print(f"  Val: {len(X_val):,} samples")
        print(f"  Test: {len(X_test):,} samples")
        print(f"  Test class distribution:")
        print(f"    Completed: {(y_test == 1).sum():,} ({(y_test == 1).mean()*100:.1f}%)")
        print(f"    Non-completed: {(y_test == 0).sum():,} ({(y_test == 0).mean()*100:.1f}%)")
        
        # 5. Apply preprocessing
        print("\n" + "=" * 80)
        print("Applying Preprocessing")
        print("=" * 80)
        X_test_array, feature_names = apply_preprocessing(X_test, metadata, scaler)
        print(f"  Final feature count: {len(feature_names)}")
        
        # Save feature names
        with open(SHAP_DIR / "transformed_feature_names.txt", 'w', encoding='utf-8') as f:
            f.write('\n'.join(feature_names))
        print(f"  Saved feature names to: transformed_feature_names.txt")
        
        # 6. Check for leakage (will be included in report, but also exit if found)
        leakage_detected = check_leakage(feature_names)
        
        # 7. Compute SHAP values
        shap_values, explainer, sample_idx = compute_shap_values(
            model, X_test_array, feature_names, metadata, max_samples=2000
        )
        X_shap = X_test_array[sample_idx]
        
        # 8. Compute feature importance
        print("\n" + "=" * 80)
        print("Computing Feature Importance")
        print("=" * 80)
        importance_df = compute_feature_importance(shap_values, feature_names)
        
        # Filter out PCA features for interpretability (they're hard to understand)
        importance_df = importance_df[~importance_df.index.str.startswith('pca_emb')]
        print(f"  Excluded PCA features for interpretability")
        print(f"  Top 5 non-PCA features:")
        for i, (feature_name, row) in enumerate(importance_df.head(5).iterrows(), 1):
            print(f"    {i}. {feature_name}: {row['mean_abs_shap']:.6f}")
        
        # Save raw top 20 (with full precision - 8 decimal places) - non-PCA only
        df_top20_raw = importance_df.head(20).reset_index()
        df_top20_raw['rank'] = range(1, len(df_top20_raw) + 1)
        df_top20_raw['feature_display'] = df_top20_raw['feature'].apply(format_feature_name)
        df_top20_raw = df_top20_raw[['rank', 'feature', 'feature_display', 'mean_abs_shap']]
        # Ensure full precision in CSV
        df_top20_raw.to_csv(SHAP_DIR / "shap_top20_raw.csv", index=False, float_format='%.8f')
        print(f"  Saved: shap_top20_raw.csv (with full precision)")
        
        # Group one-hot features
        grouped_df = group_onehot_features(importance_df)
        df_top20_grouped = grouped_df.head(20).reset_index()
        df_top20_grouped['rank'] = range(1, len(df_top20_grouped) + 1)
        df_top20_grouped['feature_display'] = df_top20_grouped['feature'].apply(format_feature_name)
        df_top20_grouped = df_top20_grouped[['rank', 'feature', 'feature_display', 'mean_abs_shap']]
        df_top20_grouped.to_csv(SHAP_DIR / "shap_top20_grouped.csv", float_format='%.8f', index=False)
        print(f"  Saved: shap_top20_grouped.csv (with full precision)")
        
        # 9. Export to Excel
        export_to_excel(importance_df, grouped_df, SHAP_DIR)
        
        # 10. Generate plots (non-PCA features only for interpretability)
        generate_shap_plots(shap_values, X_shap, feature_names, importance_df, SHAP_DIR)
        
        # For dependence plots, also filter to non-PCA features
        top_features_for_dependence = [f for f in grouped_df.head(5).index.tolist() if not f.startswith('pca_emb')]
        if top_features_for_dependence:
            generate_dependence_plots(
                shap_values, X_shap, feature_names,
                top_features_for_dependence, SHAP_DIR
            )
        else:
            print("  Skipping dependence plots: all top features are PCA (excluded for interpretability)")
        
        # 11. Generate report
        generate_report(
            metadata, importance_df, grouped_df, X_shap, shap_values,
            feature_names, SHAP_DIR
        )
        
        # Exit with error if leakage was detected
        if leakage_detected:
            print("\n" + "!" * 80)
            print("LEAKAGE DETECTED - Exiting with error status")
            print("!" * 80)
            sys.exit(1)
        
        print("\n" + "=" * 80)
        print("SHAP Analysis Complete")
        print("=" * 80)
        print(f"End time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"\nAll outputs saved to: {SHAP_DIR}")
        
    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

